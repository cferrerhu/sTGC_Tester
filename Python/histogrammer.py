#!/usr/bin/python3

import matplotlib.pyplot as plt
from os import listdir
import xlrd
import sys
import openpyxl
from openpyxl import load_workbook

# Debug options
debug = False
debugColor = False

# Get all Excel sheets for each gfz connector on a wedge
quadlist = listdir("./Results")
gfzlist = list()
for quad in quadlist:
    gfzlist.append(listdir("./Results/"+quad))
    for i in range(len(gfzlist[-1])):
        gfzlist[-1][i] = "./Results/" + quad + "/" + gfzlist[-1][i]

# swith pad vs strip
sgfzlist = list()
pgfzlist = list()
cnt = 0
for i in gfzlist:
    sgfzlist.append(list())
    pgfzlist.append(list())
    for j in i:
        if "_S_" in j: sgfzlist[-1].append(j)
        if "_P_" in j: pgfzlist[-1].append(j)
    cnt = cnt + 1
totalgfzlist = [sgfzlist, pgfzlist]

# once for the strips, once for the pads
connectorACValues = list()
padtrigger = False  # is used to distiguishe in the for-loop whether or not pads are analyzed
for gfz in totalgfzlist:
    # save them in an accessiable format
    workbooks = list()
    sheets = list()
    pyxlwbs = list()
    pyxlsheets = list()
    num = 0
    for i in gfz:
        workbooks.append(list())
        sheets.append(list())
        pyxlwbs.append(list())
        pyxlsheets.append(list())
        for j in i:
            if debug: print(j)
            try:
                workbooks[num].append(xlrd.open_workbook(j))
                pyxlwbs[num].append(load_workbook(j, data_only = True))
            except xlrd.biffh.XLRDError:
                sys.stderr.write("ERROR: Please close all excel sheets\n")
            except:
                print("Unexpected error:", sys.exc_info()[0])
                raise
            sheets[num].append(workbooks[num][-1].sheet_by_index(3))
            pyxlsheets[num].append(pyxlwbs[num][-1]['ConnectorDC'])
        num = num + 1

    # get the measurement data for all ConnectorAC data
    for i in range(len(sheets)):
        if debug: print(quadlist[i])
        data = list()
        for j in range(len(sheets[i])):
            for k in range(2,32):
                for l in range(3,13):
                    m = sheets[i][j].cell_value(k,l).split(": ")
                    colorInHex = pyxlsheets[i][j].cell(k+1,l+1).fill.start_color.index # this gives you Hexadecimal value of the color
                    if debugColor: print("Worksheet {}, ConnectorAC, row {}, column {}, color in HEX {}".format(gfz[i][j],k+1,l+1,colorInHex))
                    if debugColor: print(str(colorInHex) != '00000000')
                    if debugColor: print(str(colorInHex) != 'FFFFFFFF')
                    if(m[0] != '999' and len(m) == 2 and str(colorInHex) != '00000000' and str(colorInHex) != 'FFFFFFFF'):
                        # necessary to disregard wires if pads are regarded
                        if not padtrigger or int(m[0]) <= 127:
                            if m[1] != '1023': data.append(int(m[1]))
                            if debugColor: print(sheets[i][j].cell_value(k,l))
                            if debugColor: print(pyxlsheets[i][j].cell(k+1,l+1).value)
        if debug: print(data)
        connectorACValues.append(data)
    padtrigger = True

hatches=['','//','\\\\']
plt.rcParams['hatch.linewidth'] = 0.5
# plot histograms
plt.subplot(121)
for i in range(len(quadlist)):
    plt.hist(connectorACValues[i], label = "{} Strip Amplitude".format(quadlist[i]), bins=range(0,450,10), alpha=0.5, edgecolor='black', linewidth=1.2, hatch=hatches[i])
plt.legend(loc='upper right')
plt.ylabel('Frequency')
plt.xlabel('Amplitude [ADC counts]')
plt.title('Amplitude histogram Strips')
plt.grid(which='both')
plt.minorticks_on()
plt.xlim(xmin=0.0)
plt.ylim(ymin=0.0)

# plot histograms
plt.subplot(122)
for i in range(len(quadlist)):
    plt.hist(connectorACValues[i+len(quadlist)], label = "{} Pad Amplitude".format(quadlist[i]), bins=range(0,900,20), alpha=0.5, edgecolor='black', linewidth=1.2, hatch=hatches[i])
plt.legend(loc='upper right')
plt.ylabel('Frequency')
plt.xlabel('Amplitude [ADC counts]')
plt.title('Amplitude histogram Pads')
plt.grid(which='both')
plt.minorticks_on()
plt.xlim(xmin=0.0)
plt.ylim(ymin=0.0)

plt.show()
